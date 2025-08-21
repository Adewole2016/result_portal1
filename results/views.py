from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy

from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Semester
from .forms import SemesterForm  # Ensure you have created a SemesterForm
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Semester
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Semester
from .forms import SemesterForm
from .models import Session, Semester, Result, StudentSemesterSummary, GradingSystem, GradeScale
from courses.models import Course
from students.models import Student
from accounts.models import AuditLog
from accounts.models import CourseAllocation

from django.contrib.auth.decorators import login_required
from django.contrib import messages

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from results.forms import ResultUploadForm  # assuming you saved the form in results/forms.py
from accounts.models import CourseAllocation

from django.shortcuts import render
from django.http import HttpResponse

from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Session
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Session
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse
from .models import Session

from django.shortcuts import render

from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from django.shortcuts import get_object_or_404
from .models import Result
from courses.models import Course

from openpyxl.utils import get_column_letter

import pandas as pd
import openpyxl
import pandas as pd
import io
import io
import json
import io


@login_required
def result_dashboard(request):
    context = {
        'total_results': Result.objects.count(),
        'pending_results': Result.objects.filter(status='pending').count(),
        'approved_results': Result.objects.filter(status='approved').count(),
        'current_semester': Semester.objects.filter(is_current=True).first(),
        'current_session': Session.objects.filter(is_current=True).first(),
    }
    
    if request.user.is_lecturer:
        allocations = CourseAllocation.objects.filter(lecturer=request.user)
        context['allocations'] = allocations
        context['lecturer_results'] = Result.objects.filter(
            course__in=[alloc.course for alloc in allocations]
        ).order_by('-created_at')[:10]

    return render(request, 'results/dashboard.html', context)






@login_required
def upload_scores(request, course_id=None):
    if not hasattr(request.user, "is_lecturer") or not request.user.is_lecturer:
        messages.error(request, "Access denied. Only lecturers can upload scores.")
        return redirect("results:dashboard")

    allocations = CourseAllocation.objects.filter(lecturer=request.user)

    selected_allocation = None
    form = None

    if course_id:
        selected_allocation = get_object_or_404(
            allocations,
            course__id=course_id
        )

    if request.method == "POST" and selected_allocation:
        form = ResultUploadForm(
            request.POST,
            course=selected_allocation.course,
            semester=selected_allocation.semester,
            uploaded_by=request.user
        )
        if form.is_valid():
            form.save()
            messages.success(request, f"Result for {selected_allocation.course.code} uploaded successfully.")
            return redirect("results:dashboard")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        if selected_allocation:
            form = ResultUploadForm(
                course=selected_allocation.course,
                semester=selected_allocation.semester,
                uploaded_by=request.user
            )

    return render(request, "results/upload_scores.html", {
        "allocations": allocations,
        "selected_allocation": selected_allocation,
        "form": form,
    })




@login_required
def approve_results(request):
    """
    HOD approval of results
    """
    if not request.user.can_approve_results():
        messages.error(request, 'Access denied. You do not have permission to approve results.')
        return redirect('results:dashboard')
    
    if request.method == 'POST':
        result_ids = request.POST.getlist('result_ids')
        action = request.POST.get('action')
        
        if action == 'approve':
            Result.objects.filter(id__in=result_ids).update(
                status='approved',
                approved_by=request.user
            )
            messages.success(request, f'Approved {len(result_ids)} results.')
            
            # Log the approval
            AuditLog.objects.create(
                user=request.user,
                action='approve',
                model_name='Result',
                description=f'Approved {len(result_ids)} results',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
        elif action == 'reject':
            Result.objects.filter(id__in=result_ids).update(
                status='rejected',
                approved_by=request.user
            )
            messages.success(request, f'Rejected {len(result_ids)} results.')
    
    # Get pending results for approval
    pending_results = Result.objects.filter(status='submitted').order_by('-created_at')
    
    if request.user.is_hod:
        # HOD can only approve results for their department
        pending_results = pending_results.filter(course__department=request.user.department)
    
    paginator = Paginator(pending_results, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'pending_results': page_obj,
    }
    
    return render(request, 'results/approve_results.html', context)




@login_required
def student_results(request):
    """
    Student view of their results
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. This page is for students only.')
        return redirect('accounts:dashboard')
    
    try:
        student = request.user.student_profile
    except:
        messages.error(request, 'Student profile not found. Please contact the administrator.')
        return redirect('accounts:dashboard')
    
    # Get student's results
    results = Result.objects.filter(
        student=student,
        status='approved'
    ).order_by('-semester__session__start_year', '-semester__semester', 'course__code')
    
    # Get semester summaries
    summaries = StudentSemesterSummary.objects.filter(student=student).order_by(
        '-semester__session__start_year', '-semester__semester'
    )
    
    # Calculate overall CGPA
    if summaries.exists():
        total_credit_units = sum(s.total_credit_units for s in summaries)
        total_grade_points = sum(s.total_grade_points for s in summaries)
        overall_cgpa = total_grade_points / total_credit_units if total_credit_units > 0 else 0.0
    else:
        overall_cgpa = 0.0
    
    context = {
        'student': student,
        'results': results,
        'summaries': summaries,
        'overall_cgpa': overall_cgpa,
    }
    
    return render(request, 'results/student_results.html', context)


def get_client_ip(request):
    """
    Get client IP address from request
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def manage_sessions_semesters(request):
    return HttpResponse("Manage Sessions and Semesters Page")


def add_session(request):
    if request.method == "POST":
        name = request.POST.get('name')
        if name:
            Session.objects.create(name=name)
            return HttpResponse("Session added successfully.")
        else:
            return HttpResponse("Session name is required.", status=400)
    return render(request, 'results/add_session.html')




def edit_session(request, pk):
    session = get_object_or_404(Session, pk=pk)
    if request.method == "POST":
        name = request.POST.get('name')
        if name:
            session.name = name
            session.save()
            return redirect('results:add_session')  # Redirect to session list or add page
        else:
            return HttpResponse("Session name is required.", status=400)
    return render(request, 'results/edit_session.html', {'session': session})



def delete_session(request, pk):
    session = get_object_or_404(Session, pk=pk)
    session.delete()
    messages.success(request, "Session deleted successfully.")
    return redirect('results:add_session')  # Adjust this redirect to your sessions list page




def add_semester(request):
    if request.method == 'POST':
        form = SemesterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Semester added successfully!')
            return redirect('results:add_semester')
    else:
        form = SemesterForm()

    semesters = Semester.objects.all()
    return render(request, 'results/add_semester.html', {'form': form, 'semesters': semesters})



def edit_semester(request, pk):
    semester = get_object_or_404(Semester, pk=pk)
    if request.method == 'POST':
        form = SemesterForm(request.POST, instance=semester)
        if form.is_valid():
            form.save()
            messages.success(request, 'Semester updated successfully!')
            return redirect('results:add_semester')
    else:
        form = SemesterForm(instance=semester)

    return render(request, 'results/edit_semester.html', {'form': form, 'semester': semester})


def delete_semester(request, pk):
    semester = get_object_or_404(Semester, pk=pk)
    semester.delete()
    messages.success(request, 'Semester deleted successfully!')
    return redirect('results:add_semester')

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import GradingSystem  # Ensure you have this model defined
from .forms import GradingSystemForm  # We'll create this form

def manage_grading_system(request):
    grading_systems = GradingSystem.objects.all()
    
    if request.method == 'POST':
        form = GradingSystemForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Grading system updated successfully!")
            return redirect('results:manage_grading_system')
    else:
        form = GradingSystemForm()
    
    return render(request, 'results/manage_grading_system.html', {
        'form': form,
        'grading_systems': grading_systems
    })


from django.shortcuts import render, redirect
from django.http import HttpResponse

def manage_sessions(request):
    return HttpResponse("Manage Sessions Page (Under Construction)")

def add_session(request):
    return HttpResponse("Add Session Page (Under Construction)")

def edit_session(request, pk):
    return HttpResponse(f"Edit Session {pk} Page (Under Construction)")

def delete_session(request, pk):
    return HttpResponse(f"Delete Session {pk} Page (Under Construction)")

def manage_semesters(request):
    return HttpResponse("Manage Semesters Page (Under Construction)")

def add_semester(request):
    return HttpResponse("Add Semester Page (Under Construction)")

def edit_semester(request, pk):
    return HttpResponse(f"Edit Semester {pk} Page (Under Construction)")

def delete_semester(request, pk):
    return HttpResponse(f"Delete Semester {pk} Page (Under Construction)")

def manage_grading_system(request):
    return HttpResponse("Manage Grading System Page (Under Construction)")

def add_grading_system(request):
    return HttpResponse("Add Grading System Page (Under Construction)")

def edit_grading_system(request, pk):
    return HttpResponse(f"Edit Grading System {pk} Page (Under Construction)")



def manage_sessions_semesters(request):
    return render(request, 'results/manage_sessions_semesters.html')

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils.decorators import method_decorator
from django.views.generic import ListView, CreateView, UpdateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import json

from .models import Session, Semester, Result, StudentSemesterSummary, GradingSystem, GradeScale
from courses.models import Course
from students.models import Student
from accounts.models import AuditLog


import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

"""
def download_score_template(request, course_id):
    
    Download Excel template for score entry
    
    if not request.user.can_upload_scores():
        messages.error(request, 'Access denied. You do not have permission to download score templates.')
        return redirect('results:dashboard')
    
    course = get_object_or_404(Course, id=course_id)
    current_semester = Semester.objects.filter(is_current=True).first()
    
    if not current_semester:
        messages.error(request, 'No current semester found. Please contact the administrator.')
        return redirect('results:dashboard')
    
    # Get students registered for this course
    students = Student.objects.filter(
        level=course.level,
        department=course.department
    ).order_by('matric_no')
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{course.code} Score Sheet"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        "S/N", "Matric No", "Full Name", "CA Score (0-40)", "Exam Score (0-60)", "Total", "Grade", "Remarks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Course information
    ws.cell(row=2, column=1, value="Course:")
    ws.cell(row=2, column=2, value=f"{course.code} - {course.title}")
    ws.cell(row=3, column=1, value="Credit Units:")
    ws.cell(row=3, column=2, value=course.credit_units)
    ws.cell(row=4, column=1, value="Semester:")
    ws.cell(row=4, column=2, value=str(current_semester))
    ws.cell(row=5, column=1, value="Department:")
    ws.cell(row=5, column=2, value=course.department.name)
    
    # Student data starts from row 7
    start_row = 7
    for idx, student in enumerate(students, 1):
        row = start_row + idx - 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=student.matric_no)
        ws.cell(row=row, column=3, value=student.user.get_full_name())
        # CA and Exam score columns are left empty for input
        ws.cell(row=row, column=4, value="")  # CA Score
        ws.cell(row=row, column=5, value="")  # Exam Score
        # Formula for total
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}")
        # Grade formula (simplified)
        ws.cell(row=row, column=7, value=f'=IF(F{row}>=70,"A",IF(F{row}>=60,"B",IF(F{row}>=50,"C",IF(F{row}>=45,"D",IF(F{row}>=40,"E","F")))))')
        ws.cell(row=row, column=8, value="")  # Remarks
    
    # Adjust column widths
    column_widths = [5, 15, 25, 15, 15, 10, 8, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{course.code}_score_template.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    # Log the download
    AuditLog.objects.create(
        user=request.user,
        action='download',
        model_name='Course',
        object_id=str(course.id),
        description=f'Downloaded score template for {course.code}',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    return response
"""




"""
@login_required
def download_score_template(request, course_id):
    
    #Download Excel template for score entry
    
    if not request.user.can_upload_scores():
        messages.error(request, 'Access denied. You do not have permission to download score templates.')
        return redirect('results:dashboard')
    
    course = get_object_or_404(Course, id=course_id)
    current_semester = Semester.objects.filter(is_current=True).first()
    
    if not current_semester:
        messages.error(request, 'No current semester found. Please contact the administrator.')
        return redirect('results:dashboard')
    
    # Get students registered for this course
    students = Student.objects.filter(
        level=course.level,
        department=course.department
    ).order_by('matric_no')
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{course.code} Score Sheet"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        "S/N", "Matric No", "Full Name", "CA Score (0-40)", "Exam Score (0-60)", "Total", "Grade", "Remarks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Course information
    ws.cell(row=2, column=1, value="Course:")
    ws.cell(row=2, column=2, value=f"{course.code} - {course.title}")
    ws.cell(row=3, column=1, value="Credit Units:")
    ws.cell(row=3, column=2, value=course.credit_units)
    ws.cell(row=4, column=1, value="Semester:")
    ws.cell(row=4, column=2, value=str(current_semester))
    ws.cell(row=5, column=1, value="Department:")
    ws.cell(row=5, column=2, value=course.department.name)
    
    # Student data starts from row 7
    start_row = 7
    for idx, student in enumerate(students, 1):
        row = start_row + idx - 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=student.matric_no)
        ws.cell(row=row, column=3, value=student.user.get_full_name())
        # CA and Exam score columns are left empty for input
        ws.cell(row=row, column=4, value="")  # CA Score
        ws.cell(row=row, column=5, value="")  # Exam Score
        # Formula for total
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}")
        # Grade formula (simplified)
        ws.cell(row=row, column=7, value=f'=IF(F{row}>=70,"A",IF(F{row}>=60,"B",IF(F{row}>=50,"C",IF(F{row}>=45,"D",IF(F{row}>=40,"E","F")))))')
        ws.cell(row=row, column=8, value="")  # Remarks
    
    # Adjust column widths
    column_widths = [5, 15, 25, 15, 15, 10, 8, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{course.code}_score_template.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    # Log the download
    AuditLog.objects.create(
        user=request.user,
        action='download',
        model_name='Course',
        object_id=str(course.id),
        description=f'Downloaded score template for {course.code}',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    return response
"""

@login_required
def approve_results(request):
    """
    HOD approval of results
    """
    if not request.user.can_approve_results():
        messages.error(request, 'Access denied. You do not have permission to approve results.')
        return redirect('results:dashboard')
    
    if request.method == 'POST':
        result_ids = request.POST.getlist('result_ids')
        action = request.POST.get('action')
        
        if action == 'approve':
            Result.objects.filter(id__in=result_ids).update(
                status='approved',
                approved_by=request.user
            )
            messages.success(request, f'Approved {len(result_ids)} results.')
            
            # Log the approval
            AuditLog.objects.create(
                user=request.user,
                action='approve',
                model_name='Result',
                description=f'Approved {len(result_ids)} results',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
        elif action == 'reject':
            Result.objects.filter(id__in=result_ids).update(
                status='rejected',
                approved_by=request.user
            )
            messages.success(request, f'Rejected {len(result_ids)} results.')
    
    # Get pending results for approval
    pending_results = Result.objects.filter(status='submitted').order_by('-created_at')
    
    if request.user.is_hod:
        # HOD can only approve results for their department
        pending_results = pending_results.filter(course__department=request.user.department)
    
    paginator = Paginator(pending_results, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'pending_results': page_obj,
    }
    
    return render(request, 'results/approve_results.html', context)

@login_required
def student_results(request):
    """
    Student view of their results
    """
    if not request.user.is_student:
        messages.error(request, 'Access denied. This page is for students only.')
        return redirect('accounts:dashboard')
    
    try:
        student = request.user.student_profile
    except:
        messages.error(request, 'Student profile not found. Please contact the administrator.')
        return redirect('accounts:dashboard')
    
    # Get student's results
    results = Result.objects.filter(
        student=student,
        status='approved'
    ).order_by('-semester__session__start_year', '-semester__semester', 'course__code')
    
    # Get semester summaries
    summaries = StudentSemesterSummary.objects.filter(student=student).order_by(
        '-semester__session__start_year', '-semester__semester'
    )
    
    # Calculate overall CGPA
    if summaries.exists():
        total_credit_units = sum(s.total_credit_units for s in summaries)
        total_grade_points = sum(s.total_grade_points for s in summaries)
        overall_cgpa = total_grade_points / total_credit_units if total_credit_units > 0 else 0.0
    else:
        overall_cgpa = 0.0
    
    context = {
        'student': student,
        'results': results,
        'summaries': summaries,
        'overall_cgpa': overall_cgpa,
    }
    
    return render(request, 'results/student_results.html', context)

def get_client_ip(request):
    """
    Get client IP address from request
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip




from .forms import BroadsheetFilterForm
from django.db.models import Sum, F
@login_required
def broadsheet_view(request):
    if not (request.user.is_admin or request.user.is_hod):
        messages.error(request, "Access denied. Admin or HOD privileges required.")
        return redirect("results:dashboard")

    form = BroadsheetFilterForm(request.GET or None)
    broadsheet_data = []
    selected_department = None
    selected_session = None
    selected_semester = None
    selected_level = None
    all_courses = []
    course_units = []

    total_students = total_passed = total_failed = 0
    percent_passed = percent_failed = 0
    show_prev = False  # default to False

    if form.is_valid():
        selected_department = form.cleaned_data["department"]
        selected_session = form.cleaned_data["session"]
        selected_semester = form.cleaned_data["semester"]
        selected_level = form.cleaned_data.get("level")

        # Fetch all semesters ordered chronologically
        all_semesters = Semester.objects.order_by('session__start_year', 'semester')
        semesters_list = list(all_semesters)

        # Get approved results for selected semester
        results = Result.objects.filter(
            student__department=selected_department,
            semester=selected_semester,
            status="approved"
        ).order_by("student__matric_no", "course__code")

        # Get courses for selected semester
        course_ids = results.values_list("course_id", flat=True).distinct()
        all_courses = Course.objects.filter(
            id__in=course_ids,
            department=selected_department,
            semester=selected_semester.semester
        ).order_by("code")
        course_units = [course.credit_units for course in all_courses]

        # Build student -> scores dictionary
        students_results = {}
        for r in results:
            student = r.student
            entry = students_results.setdefault(student, {"student": student, "course_scores": {}})
            entry["course_scores"][r.course.code] = r.total_score

        # Function to calculate grade point
        def calculate_gp(score):
            if score >= 75: return 4.0
            elif score >= 70: return 3.5
            elif score >= 65: return 3.25
            elif score >= 60: return 3.0
            elif score >= 55: return 2.75
            elif score >= 50: return 2.5
            elif score >= 45: return 2.25
            elif score >= 40: return 2.0
            else: return 0

        # Initialize cumulative dictionary
        student_cumulative = {}

        # Find index of selected semester
        try:
            selected_index = semesters_list.index(selected_semester)
        except ValueError:
            selected_index = 0

        # Process each semester up to the selected semester for cumulative calculation
        for sem in semesters_list[:selected_index + 1]:
            sem_results = Result.objects.filter(
                student__department=selected_department,
                semester=sem,
                status="approved"
            )

            sem_students = {}
            for r in sem_results:
                s = r.student
                entry = sem_students.setdefault(s, {"student": s, "course_scores": {}})
                entry["course_scores"][r.course.code] = r.total_score

            # Get courses for this semester
            course_ids_sem = sem_results.values_list("course_id", flat=True).distinct()
            sem_courses = Course.objects.filter(
                id__in=course_ids_sem,
                department=selected_department,
                semester=sem.semester
            )

            for student, entry in sem_students.items():
                total_credit_units = 0
                total_grade_points = 0.0
                failed_courses_current = []

                for course in sem_courses:
                    score = entry["course_scores"].get(course.code, 0)
                    total_credit_units += course.credit_units
                    gp = calculate_gp(score)
                    total_grade_points += gp * course.credit_units
                    if score < 40:
                        failed_courses_current.append(course.code)

                gpa = round(total_grade_points / total_credit_units, 2) if total_credit_units else 0

                # Previous cumulative
                prev = student_cumulative.get(student)
                if prev:
                    prev_tnu = prev["ctnu"]
                    prev_tcp = prev["ctcp"]
                    prev_gpa = prev["cgpa"]
                    cumulative_failed_courses = list(set(prev["failed_courses"] + failed_courses_current))
                else:
                    prev_tnu = prev_tcp = prev_gpa = 0
                    cumulative_failed_courses = failed_courses_current.copy()

                # Cumulative for this semester
                ctnu = total_credit_units + prev_tnu
                ctcp = total_grade_points + prev_tcp
                cgpa = round(ctcp / ctnu, 2) if ctnu else 0

                remark = "PASSED" if not cumulative_failed_courses else "FAILED"

                # Append to broadsheet only if this is selected semester
                if sem == selected_semester:
                    show_prev = prev_tnu > 0
                    broadsheet_data.append({
                        "student": student,
                        "course_scores": entry["course_scores"],
                        "tnu": total_credit_units,
                        "tcp": total_grade_points,
                        "gpa": gpa,
                        "prev_tnu": prev_tnu,
                        "prev_tcp": prev_tcp,
                        "prev_gpa": prev_gpa,
                        "ctnu": ctnu,
                        "ctcp": ctcp,
                        "cgpa": cgpa,
                        "failed_courses": cumulative_failed_courses,
                        "remark": remark
                    })

                # Update cumulative dictionary
                student_cumulative[student] = {
                    "ctnu": ctnu,
                    "ctcp": ctcp,
                    "cgpa": cgpa,
                    "failed_courses": cumulative_failed_courses
                }

        # Calculate total students, passed, failed
        total_students = len(broadsheet_data)
        total_passed = sum(1 for r in broadsheet_data if r["remark"] == "PASSED")
        total_failed = total_students - total_passed
        if total_students > 0:
            percent_passed = round((total_passed / total_students) * 100, 2)
            percent_failed = round((total_failed / total_students) * 100, 2)
            # Sort broadsheet data by student's matric number
    broadsheet_data.sort(key=lambda x: x["student"].matric_no)


    return render(request, "results/broadsheet1.html", {
        "form": form,
        "broadsheet_data": broadsheet_data,
        "selected_department": selected_department,
        "selected_session": selected_session,
        "selected_semester": selected_semester,
        "selected_level": selected_level,
        "all_courses": all_courses,
        "course_units": course_units,
        "total_students": total_students,
        "total_passed": total_passed,
        "total_failed": total_failed,
        "percent_passed": percent_passed,
        "percent_failed": percent_failed,
        "show_prev": show_prev
    })

"""
@login_required
def broadsheet_view(request):
    if not (request.user.is_admin or request.user.is_hod):
        messages.error(request, "Access denied. Admin or HOD privileges required.")
        return redirect("results:dashboard")

    form = BroadsheetFilterForm(request.GET or None)
    broadsheet_data = []
    selected_department = None
    selected_session = None
    selected_semester = None
    all_courses = []
    course_units = []

    if form.is_valid():
        selected_department = form.cleaned_data["department"]
        selected_session = form.cleaned_data["session"]      # Session instance
        selected_semester = form.cleaned_data["semester"]    # Semester instance

        # Get approved results for the department, session, and semester
        results = Result.objects.filter(
            student__department=selected_department,
            semester=selected_semester,
            semester__session=selected_session,
            status="approved"
        ).order_by("student__matric_no", "course__code")

        # Get only the courses that appear in these results
        course_ids = results.values_list("course_id", flat=True).distinct()
        all_courses = Course.objects.filter(
            id__in=course_ids,
            department=selected_department,
            semester=selected_semester.semester  # "first" or "second"
        ).order_by("code")
        course_units = [course.credit_units for course in all_courses]

        # Build dictionary of student → course scores
        students_results = {}
        for r in results:
            student = r.student
            entry = students_results.setdefault(student, {"student": student, "course_scores": {}})
            entry["course_scores"][r.course.code] = r.total_score

        # Fill missing scores with None, compute totals & GPA ignoring None
        for student, entry in students_results.items():
            total_credit_units = 0
            total_grade_points = 0.0
            failed_courses = []

            for course in all_courses:
                score = entry["course_scores"].get(course.code)

                # Unattempted or missing score
                if score is None or score == 0:
                    entry["course_scores"][course.code] = "-"  # show dash
                    failed_courses.append(course.code)        # count as failed
                    continue  # skip in GPA computation

                entry["course_scores"][course.code] = score
                total_credit_units += course.credit_units

                # Grade point calculation
                if score >= 75:
                    gp = 4.00
                elif score >= 70:
                    gp = 3.5
                elif score >= 65:
                    gp = 3.25
                elif score >= 60:
                    gp = 3.00
                elif score >= 55:
                    gp = 2.75
                elif score >= 50:
                    gp = 2.5
                elif score >= 45:
                    gp = 2.25
                elif score >= 40:
                    gp = 2.00
                else:
                    gp = 0

                total_grade_points += gp * course.credit_units

                if score < 40:
                    failed_courses.append(course.code)

            gpa = round(total_grade_points / total_credit_units, 2) if total_credit_units else 0

            broadsheet_data.append({
                "student": student,
                "course_scores": entry["course_scores"],
                "total_credit_units": total_credit_units,
                "total_grade_points": total_grade_points,
                "gpa": gpa,
                "tnu": total_credit_units,
                "tcp": total_grade_points,
                "remark": "PASSED" if not failed_courses else "FAILED",
                "failed_courses": failed_courses,
                "failed_courses_str": ", ".join(failed_courses) if failed_courses else "-"
            })

    return render(request, "results/broadsheet.html", {
        "form": form,
        "broadsheet_data": broadsheet_data,
        "selected_department": selected_department,
        "selected_session": selected_session,
        "selected_semester": selected_semester,
        "all_courses": all_courses,
        "course_units": course_units
    })

"""































from django.shortcuts import render, redirect
from django.contrib import messages
from .utils import process_uploaded_scores

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .utils import process_uploaded_scores
from courses.models import Course
from .models import Semester

import openpyxl
from django.core.exceptions import ObjectDoesNotExist
from students.models import Student 
from results.models import Result  # adjust names to your project

"""

def process_uploaded_scores(file_obj, course=None, semester=None, uploaded_by=None):
    
    Process uploaded scores from Excel and return summary.
    

    wb = openpyxl.load_workbook(file_obj)
    sheet = wb.active

    row_errors = []
    missing_students = []
    processed_count = 0

    # If semester not provided, set it to current semester
    if semester is None:
        from .models import Semester  # adjust import
        semester = Semester.get_current()

    for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Skip header rows and repeated headers
        if row_num == 1:
            continue
        if not row or not row[0]:
            continue
        if str(row[0]).strip().lower() == "matric no":
            continue

        matric_no = str(row[0]).strip()
        # Read CA & Exam safely
        try:
            ca_score = float(row[2]) if row[2] is not None else None
            exam_score = float(row[3]) if row[3] is not None else None
        except (ValueError, TypeError):
            row_errors.append(f"Row {row_num}: invalid score format for '{matric_no}'")
            continue

        if ca_score is None or exam_score is None:
            row_errors.append(f"Row {row_num}: missing CA or Exam score for '{matric_no}'")
            continue

        if not (0 <= ca_score <= 40) or not (0 <= exam_score <= 60):
            row_errors.append(f"Row {row_num}: invalid score values for '{matric_no}'")
            continue

        # Get student
        try:
            student = Student.objects.get(matric_no=matric_no)
        except ObjectDoesNotExist:
            missing_students.append(matric_no)
            continue

        # Save result
        Result.objects.update_or_create(
            student=student,
            course=course,
            semester=semester,
            defaults={
                'ca_score': ca_score,
                'exam_score': exam_score,
                'total_score': ca_score + exam_score,
                'uploaded_by': uploaded_by
            }
        )

        processed_count += 1

    return {
        'processed': processed_count,
        'course': course,
        'row_errors': row_errors,
        'missing': missing_students
    }
"""






"""
def process_uploaded_scores(request, course_id):
    if not request.user.can_upload_scores():
        messages.error(request, 'Access denied. You do not have permission to upload scores.')
        return redirect('results:dashboard')
    
    course = get_object_or_404(Course, id=course_id)
    current_semester = Semester.objects.filter(is_current=True).first()
    
    if not current_semester:
        messages.error(request, 'No current semester found. Please contact the administrator.')
        return redirect('results:dashboard')
    
    if request.method == 'POST' and request.FILES.get('score_file'):
        score_file = request.FILES['score_file']
        
        try:
            wb = openpyxl.load_workbook(score_file)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Failed to read uploaded file: {e}")
            return redirect('results:upload_scores', course_id=course.id)
        
        errors = []
        start_row = 7  # Based on your template structure
        
        for idx, row in enumerate(ws.iter_rows(min_row=start_row), start=1):
            matric_no_cell = row[1]  # Column 2: Matric No
            ca_score_cell = row[3]   # Column 4: CA Score
            exam_score_cell = row[4] # Column 5: Exam Score
            
            matric_no = str(matric_no_cell.value).strip() if matric_no_cell.value else ''
            
            if not matric_no:
                errors.append(f"Row {start_row + idx - 1}: Missing matriculation number.")
                continue
            
            try:
                student = Student.objects.get(matric_no__iexact=matric_no)
            except Student.DoesNotExist:
                errors.append(f"Row {start_row + idx - 1}: Student with matric no '{matric_no}' not found.")
                continue
            
            # Validate CA score
            try:
                ca_score = float(ca_score_cell.value)
                if not (0 <= ca_score <= 40):
                    raise ValueError("CA score out of range")
            except Exception:
                errors.append(f"Row {start_row + idx - 1}: Invalid CA score for '{matric_no}'. Must be between 0 and 40.")
                continue
            
            # Validate Exam score
            try:
                exam_score = float(exam_score_cell.value)
                if not (0 <= exam_score <= 60):
                    raise ValueError("Exam score out of range")
            except Exception:
                errors.append(f"Row {start_row + idx - 1}: Invalid Exam score for '{matric_no}'. Must be between 0 and 60.")
                continue
            
            total_score = ca_score + exam_score
            
            # Compute grade - same as your formula
            if total_score >= 70:
                grade = 'A'
            elif total_score >= 60:
                grade = 'B'
            elif total_score >= 50:
                grade = 'C'
            elif total_score >= 45:
                grade = 'D'
            elif total_score >= 40:
                grade = 'E'
            else:
                grade = 'F'
            
            # Save or update score record in your model (example Score model)
            Score.objects.update_or_create(
                student=student,
                course=course,
                semester=current_semester,
                defaults={
                    'ca_score': ca_score,
                    'exam_score': exam_score,
                    'total_score': total_score,
                    'grade': grade,
                }
            )
        
        if errors:
            messages.warning(request, "Upload completed with some issues:\n" + "\n".join(errors))
        else:
            messages.success(request, "Scores uploaded successfully.")
        
        return redirect('results:upload_scores', course_id=course.id)
    
    else:
        messages.error(request, "No file uploaded.")
        return redirect('results:upload_scores', course_id=course.id)
"""



from accounts.models import CourseAllocation


def get_courses_for_user(user):
    allocations = CourseAllocation.objects.filter(lecturer=user).select_related('course', 'course__department')
    return [alloc.course for alloc in allocations]


import pandas as pd
from students.models import Student
from results.models import Result, GradingSystem
from django.core.exceptions import ObjectDoesNotExist

import pandas as pd
from students.models import Student
from results.models import Result, GradingSystem
import pandas as pd
import os
from students.models import Student
from results.models import Result, GradingSystem

import os
import pandas as pd
from django.core.exceptions import ObjectDoesNotExist

import os
import pandas as pd
from django.core.exceptions import ObjectDoesNotExist

import os
import pandas as pd
from django.core.exceptions import ObjectDoesNotExist

import pandas as pd
import os


@login_required


def process_uploaded_scores(request, course_id):
    if not request.user.can_upload_scores():
        messages.error(request, 'Access denied. You do not have permission to upload scores.')
        return redirect('results:dashboard')
    
    course = get_object_or_404(Course, id=course_id)
    current_semester = Semester.objects.filter(is_current=True).first()
    
    if not current_semester:
        messages.error(request, 'No current semester found. Please contact the administrator.')
        return redirect('results:dashboard')
    
    if request.method == 'POST' and request.FILES.get('score_file'):
        score_file = request.FILES['score_file']
        
        try:
            wb = openpyxl.load_workbook(score_file)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Failed to read uploaded file: {e}")
            return redirect('results:upload_scores', course_id=course.id)
        
        errors = []
        start_row = 7  # Based on your template structure
        
        for idx, row in enumerate(ws.iter_rows(min_row=start_row), start=1):
            matric_no_cell = row[1]  # Column 2: Matric No
            ca_score_cell = row[3]   # Column 4: CA Score
            exam_score_cell = row[4] # Column 5: Exam Score
            
            matric_no = str(matric_no_cell.value).strip() if matric_no_cell.value else ''
            
            if not matric_no:
                errors.append(f"Row {start_row + idx - 1}: Missing matriculation number.")
                continue
            
            try:
                student = Student.objects.get(matric_no__iexact=matric_no)
            except Student.DoesNotExist:
                errors.append(f"Row {start_row + idx - 1}: Student with matric no '{matric_no}' not found.")
                continue
            
            # Validate CA score
            try:
                ca_score = float(ca_score_cell.value)
                if not (0 <= ca_score <= 40):
                    raise ValueError("CA score out of range")
            except Exception:
                errors.append(f"Row {start_row + idx - 1}: Invalid CA score for '{matric_no}'. Must be between 0 and 40.")
                continue
            
            # Validate Exam score
            try:
                exam_score = float(exam_score_cell.value)
                if not (0 <= exam_score <= 60):
                    raise ValueError("Exam score out of range")
            except Exception:
                errors.append(f"Row {start_row + idx - 1}: Invalid Exam score for '{matric_no}'. Must be between 0 and 60.")
                continue
            
            total_score = ca_score + exam_score
            
            # Compute grade - same as your formula
            if total_score >= 70:
                grade = 'A'
            elif total_score >= 60:
                grade = 'B'
            elif total_score >= 50:
                grade = 'C'
            elif total_score >= 45:
                grade = 'D'
            elif total_score >= 40:
                grade = 'E'
            else:
                grade = 'F'
            
            # Save or update score record in your model (example Score model)
            Score.objects.update_or_create(
                student=student,
                course=course,
                semester=current_semester,
                defaults={
                    'ca_score': ca_score,
                    'exam_score': exam_score,
                    'total_score': total_score,
                    'grade': grade,
                }
            )
        
        if errors:
            messages.warning(request, "Upload completed with some issues:\n" + "\n".join(errors))
        else:
            messages.success(request, "Scores uploaded successfully.")
        
        return redirect('results:upload_scores', course_id=course.id)
    
    else:
        messages.error(request, "No file uploaded.")
        return redirect('results:upload_scores', course_id=course.id)







from django.shortcuts import render, redirect
from .forms import ScoreUploadForm
from .utils import process_uploaded_scores  # where your function is defined
from courses.models import Course
from results.models import Session, Semester

def upload_scores_view(request):
    if request.method == 'POST':
        form = ScoreUploadForm(request.POST, request.FILES)
        if form.is_valid():
            score_file = request.FILES['score_file']

            # Get course, session, semester from the form or request
            course_id = form.cleaned_data['course']
            session_id = form.cleaned_data['session']
            semester_id = form.cleaned_data['semester']

            course = Course.objects.get(id=course_id)
            session = Session.objects.get(id=session_id)
            semester = Semester.objects.get(id=semester_id)

            try:
                process_uploaded_scores(score_file, course, session, semester)
                return redirect('success_page')  # Replace with your success URL
            except Exception as e:
                form.add_error(None, f"Upload failed: {e}")
    else:
        form = ScoreUploadForm()

    return render(request, 'upload_scores.html', {'form': form})

from django.shortcuts import render, get_object_or_404
from courses.models import Course
from results.models import Result  # adjust if your Result model lives elsewhere

def view_results(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    
    # Get the latest semester with results for this course
    latest_result = Result.objects.filter(course=course).order_by('-semester__id').first()
    
    if latest_result:
        semester = latest_result.semester
        results = Result.objects.filter(course=course, semester=semester)
    else:
        semester = None
        results = Result.objects.none()
    
    return render(request, 'results/view_results.html', {
        'course': course,
        'results': results,
        'semester': semester,
    })


from django.shortcuts import render, get_object_or_404
from courses.models import Course

# results/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages

from courses.models import Course
from accounts.models import CourseAllocation
from results.models import Result
from .utils import process_uploaded_scores, get_current_semester

@login_required
def upload_results(request, course_id):
    """
    Upload scores for a specific course (button from allocations list should link here).
    """
    course = get_object_or_404(Course, id=course_id)

    # optional: ensure lecturer is allocated to this course
    if hasattr(request.user, 'role') and request.user.role == 'lecturer':
        allowed = CourseAllocation.objects.filter(course=course, lecturer=request.user).exists()
        if not allowed:
            messages.error(request, "You are not allocated to this course.")
            return redirect('results:upload_scores')

    if request.method == 'POST':
        uploaded_file = request.FILES.get('score_file')
        if not uploaded_file:
            messages.error(request, "Please select a file to upload.")
            return redirect('results:upload_results', course_id=course.id)

        try:
            semester = get_current_semester()
            summary = process_uploaded_scores(
                uploaded_file,
                course=course,
                semester=semester,
                uploaded_by=request.user
            )

            # Create messages
            if summary['processed'] > 0:
                messages.success(request, f"{summary['processed']} rows processed ({summary['created']} created, {summary['updated']} updated).")

            if summary['missing']:
                messages.warning(request, f"{len(summary['missing'])} matric numbers not found: {', '.join(summary['missing'][:10])}" +
                                 (f"... and {len(summary['missing'])-10} more." if len(summary['missing'])>10 else ""))

            if summary['row_errors']:
                # show up to 10 errors
                errs = summary['row_errors'][:10]
                messages.error(request, "Some rows had errors:\n" + "\n".join(errs))

            return redirect('results:view_results', course_id=course.id)

        except Exception as e:
            messages.error(request, f"Failed to process file: {e}")
            return redirect('results:upload_results', course_id=course.id)

    # GET -> render upload form
    return render(request, 'results/upload_results.html', {'course': course})


@login_required
def view_results(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    # Only show approved/submitted depending on policy; here all statuses
    results = Result.objects.filter(course=course).order_by('student__matric_no')
    return render(request, 'results/view_results.html', {'course': course, 'results': results})



def download_excel(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    results = Result.objects.filter(course=course)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{course.code} Results"

    headers = ['S/N', 'Matric No', 'Full Name', 'CA Score', 'Exam Score', 'Total Score', 'Remarks']
    ws.append(headers)

    for i, r in enumerate(results, start=1):
        remarks = "Passed" if r.total_score >= 40 else "Failed"
        ws.append([
            i,
            r.student.matric_no,
            r.student.user.get_full_name(),
            int(r.ca_score),
            int(r.exam_score),
            int(r.total_score),
            remarks,
        ])

    # Adjust column widths for better readability
    column_widths = {
        1: 6,    # S/N
        2: 15,   # Matric No
        3: 30,   # Full Name
        4: 10,   # CA Score
        5: 10,   # Exam Score
        6: 12,   # Total Score
        7: 10,   # Remarks
    }
    for col_num, width in column_widths.items():
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={course.code}_results.xlsx'
    wb.save(response)
    return response





def download_pdf(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    results = Result.objects.filter(course=course)

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=landscape(letter))  # Landscape for more width

    width, height = landscape(letter)
    margin_left = 40
    y = height - 50

    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(margin_left, y, f"Results for {course.code} - {course.title}")

    # Headers
    y -= 30
    p.setFont("Helvetica-Bold", 12)
    headers = ['S/N', 'Matric No', 'Full Name', 'CA', 'Exam', 'Total', 'Remarks']
    x_positions = [margin_left, margin_left + 60, margin_left + 160, margin_left + 400, margin_left + 460, margin_left + 520, margin_left + 580]

    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y, header)

    # Rows
    y -= 20
    p.setFont("Helvetica", 10)
    line_height = 18
    for i, r in enumerate(results, start=1):
        if y < 50:
            p.showPage()
            y = height - 50
            # Redraw headers on new page
            p.setFont("Helvetica-Bold", 12)
            for j, header in enumerate(headers):
                p.drawString(x_positions[j], y, header)
            y -= 20
            p.setFont("Helvetica", 10)

        remarks = "Passed" if r.total_score >= 40 else "Failed"

        p.drawString(x_positions[0], y, str(i))
        p.drawString(x_positions[1], y, r.student.matric_no)
        p.drawString(x_positions[2], y, r.student.user.get_full_name())
        p.drawString(x_positions[3], y, str(int(r.ca_score)))
        p.drawString(x_positions[4], y, str(int(r.exam_score)))
        p.drawString(x_positions[5], y, str(int(r.total_score)))
        p.drawString(x_positions[6], y, remarks)

        y -= line_height

    p.save()
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')




def download_broadsheet_excel(request):
    department_id = request.GET.get('department')
    session_id = request.GET.get('session')
    semester_id = request.GET.get('semester')

    # ✅ Corrected filter
    results = Result.objects.filter(
        student__department_id=department_id,
        semester__session_id=session_id,
        semester_id=semester_id
    )

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Broadsheet"

    # Example header row
    ws.append(["Matric No", "Full Name", "Course 1", "Course 2", "GPA"])

    # Example loop — replace with your real fields
    for res in results:
        ws.append([
            res.student.matric_no,
            res.student.user.get_full_name(),
            res.course.code,
            res.total_score,
            res.grade_point
        ])


def download_score_template(request, course_id):
    if not request.user.can_upload_scores():
        messages.error(request, 'Access denied. You do not have permission to download score templates.')
        return redirect('results:dashboard')
    
    course = get_object_or_404(Course, id=course_id)
    current_semester = Semester.objects.filter(is_current=True).first()
    
    if not current_semester:
        messages.error(request, 'No current semester found. Please contact the administrator.')
        return redirect('results:dashboard')
    
    students = Student.objects.filter(
        level=course.level,
        department=course.department
    ).order_by('matric_no')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{course.code} Score Sheet"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = [
        "S/N", "Matric No", "Full Name", "CA Score (0-40)", "Exam Score (0-60)", "Total", "Grade", "Remarks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Course info
    ws.cell(row=2, column=1, value="Course:")
    ws.cell(row=2, column=2, value=f"{course.code} - {course.title}")
    ws.cell(row=3, column=1, value="Credit Units:")
    ws.cell(row=3, column=2, value=course.credit_units)
    ws.cell(row=4, column=1, value="Semester:")
    ws.cell(row=4, column=2, value=str(current_semester))
    ws.cell(row=5, column=1, value="Department:")
    ws.cell(row=5, column=2, value=course.department.name)
    
    # Hidden course ID
    ws.cell(row=1, column=26, value=f"COURSE_ID:{course.id}")

    start_row = 7
    for idx, student in enumerate(students, 1):
        row = start_row + idx - 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=student.matric_no)
        full_name = student.user.get_full_name()  # ✅ Use User model's method
        ws.cell(row=row, column=3, value=full_name)
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}")
        ws.cell(row=row, column=7, value=f'=IF(F{row}>=70,"A",IF(F{row}>=60,"B",IF(F{row}>=50,"C",IF(F{row}>=45,"D",IF(F{row}>=40,"E","F")))))')
        ws.cell(row=row, column=8, value="")

    column_widths = [5, 15, 25, 15, 15, 10, 8, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{course.code}_score_template.xlsx"'
    wb.save(response)
    
    AuditLog.objects.create(
        user=request.user,
        action='download',
        model_name='Course',
        object_id=str(course.id),
        description=f'Downloaded score template for {course.code}',
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    return response











