import openpyxl
from django.contrib import messages
from django.shortcuts import render, redirect
from accounts.models import User
from courses.models import Department
from .models import Student

import openpyxl
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import Student, Department
from accounts.models import User

def bulk_upload_students(request):
    if request.method == "POST":
        excel_file = request.FILES.get("file")
        if not excel_file:
            messages.error(request, "No file uploaded.")
            return redirect('bulk_upload_students')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip completely empty rows
                if not any(row):
                    continue

                # Take only the first 5 columns, pad if missing
                data = list(row[:5]) + [None] * (5 - len(row))
                matric_no, full_name, dept_name, level, gender = data

                # Check for required fields
                if not all([matric_no, full_name, dept_name, level, gender]):
                    messages.warning(request, f"Skipping row with missing data: {row}")
                    continue

                # Validate department
                try:
                    department = Department.objects.get(name=dept_name)
                except Department.DoesNotExist:
                    messages.error(request, f"Department '{dept_name}' does not exist. Add it first.")
                    return redirect('bulk_upload_students')

                # Split full name into first and last name
                name_parts = str(full_name).strip().split(" ", 1)
                first_name = name_parts[0]
                last_name = name_parts[1] if len(name_parts) > 1 else ""

                # Create or get the user
                user, created = User.objects.get_or_create(
                    username=matric_no,
                    defaults={
                        "first_name": first_name,
                        "last_name": last_name,
                        "email": f"{matric_no}@school.com",
                        "role": "student",
                    },
                )

                # Create student record
                if not Student.objects.filter(matric_no=matric_no).exists():
                    Student.objects.create(
                        user=user,
                        matric_no=matric_no,
                        department=department,
                        level=level,
                        gender=gender,
                    )

            messages.success(request, "Students uploaded successfully!")
            return redirect('bulk_upload_students')

        except Exception as e:
            messages.error(request, f"Error reading Excel file: {e}")
            return redirect('bulk_upload_students')

    return render(request, "students/bulk_upload_students.html")




import openpyxl
from django.http import HttpResponse
from .models import Student

def export_students(request):
    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # Add header row
    headers = ["Matric No", "First Name", "Last Name", "Department", "Level", "Gender", "DOB", "Phone", "Address"]
    ws.append(headers)

    # Add student data rows
    for student in Student.objects.select_related("user", "department").all():
        ws.append([
            student.matric_no,
            student.user.first_name,
            student.user.last_name,
            student.department.name if student.department else "",
            student.level,
            student.get_gender_display(),
            student.date_of_birth.strftime("%Y-%m-%d") if student.date_of_birth else "",
            student.phone_number,
            student.address,
        ])

    # Create response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=students_export.xlsx'
    wb.save(response)
    return response

